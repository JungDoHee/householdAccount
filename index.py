import os
import pandas as pd
from datetime import date
from openpyxl.styles import PatternFill, Alignment, Border, Side
import re
import unicodedata

from Config import document
from Excel import excel_header as xl_header
from File import open_file as open
from Format import basic_format as format

# 산출물 폴더 경로 유무 확인 및 생성
if os.path.isdir(document.EXPORT_DOCUMENT) == False :
    os.mkdir(document.EXPORT_DOCUMENT)

# 가계부를 불러온다
files = os.listdir(document.ACCOUNT_DOCUMENT)

# 가계부의 데이터를 읽어온다
excel_data = pd.DataFrame()

dictBorder = {
    'left' : 'thin',
    'right' : 'thin',
    'top' : 'thin',
    'bottom' : 'thin'
}

def cell_color(sheet, column_len, column_no, color, type='solid') :
    for col_num in range(column_len) :
            cell = sheet.cell(row=1, column=column_no+col_num)
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type=type)

def cell_align(sheet, column_list, align_type) :
    for col_num in column_list : 
        for row in range(2, sheet.max_row+1) : 
            sheet.cell(row=row, column=col_num).alignment = Alignment(horizontal=align_type, vertical='center')

def cell_line(sheet, dicBorder) :
    thin_border = Border(
            left=Side(style=dicBorder['left']),
            right=Side(style=dicBorder['right']),
            top=Side(style=dicBorder['top']),
            bottom=Side(style=dicBorder['bottom'])
        )

    for row in sheet.iter_rows() :
        for cell in row:
            cell.border = thin_border

def get_display_width(text):
    width = 0
    for char in str(text) : 
        if unicodedata.east_asian_width(char) in ('F', 'W'):
            width += 2
        else : 
            width += 1
    return width

def cell_auto_width(sheet):
    column_list = sheet.columns
    for col in column_list :
        max_length = 0
        for cell in col:
            try:
                tmp_length = get_display_width(cell.value)
                if tmp_length > max_length:
                    max_length = tmp_length
            except:
                pass

            cell_cor = cell.coordinate
            cell_column = re.sub(r'[0-9]', '', cell_cor)

            worksheet.column_dimensions[cell_column].width = max_length
        

for file_name in files:
    card_type = file_name.split("_")[0].lower()
    fileData = pd.DataFrame(open.openFile(document.ACCOUNT_DOCUMENT+'/'+file_name, xl_header.getStartHeader(card_type)))
    if card_type == 'hyundaicard' : 
        fileData['은행/카드/증권사'] = '현대카드'
        fileData['수입금액'] = 0
    
    fileData = fileData.iloc[:, xl_header.getHeaderList(card_type)]
    fileData.columns = ['결제일', '은행/카드/증권사', '사용처', '지출금액', '수입금액']
    fileData['결제일'] = fileData['결제일'].apply(format.normalizeDateFomat)
    fileData['결제일'] = pd.to_datetime(fileData['결제일'], errors='coerce')

    # 카드 이용일 경우 지출 금액에 저장
    # 카드사 엑셀인 경우 '취소'인 경우 = 수입금액
    if card_type == 'kb' : 
        if '취소' in fileData['수입금액'] :
            fileData['수입금액'] = fileData['지출금액']
        else : 
            fileData['수입금액'] = 0
    elif card_type == 'lt' and (fileData['수입금액'] < 0).any() : 
        fileData['수입금액'] = (-1) * fileData['수입금액']

    # TODO (확장성)
    # 카드사 엑셀 뿐만이 아니라 은행사 엑셀을 넣었을 때 지출, 수입 금액 분리 방식 고려

    # 데이터 이어붙이기
    excel_data = pd.concat([excel_data, fileData], ignore_index=True)

# 월별 정렬 시작
excel_data['결제일'] = pd.to_datetime(excel_data['결제일'])
excel_data = excel_data.sort_values(by='결제일')
excel_data['월'] = excel_data['결제일'].dt.to_period('M').astype(str)

# 산출물 파일 만들기 (월별 시트 분리)
output_file_name = '{0}-{1}_{2}_가계부.xlsx'.format(excel_data['월'].iloc[0], excel_data['월'].iloc[-1], date.today())
with pd.ExcelWriter(document.EXPORT_DOCUMENT+'/'+output_file_name, engine='openpyxl') as writer : 
    for month, group in excel_data.groupby('월') : 
        # 은행/카드/증권사별 소계 계산
        group['소계'] = group['수입금액'].fillna(0) - group['지출금액'].fillna(0)
        mid_data = (
            group.groupby('은행/카드/증권사', as_index=False).agg({
                '지출금액' : 'sum',
                '수입금액' : 'sum',
                '소계' : 'sum'
            })
        )
        mid_data.columns = ['은행/카드/증권사', '지출소계', '수입소계', '소계']
        mid_data.sort_values(by='은행/카드/증권사')

        # 항목별 소계 계산
        end_data = (
            group.groupby('사용처', as_index=False).agg({
                '지출금액' : 'sum',
                '수입금액' : 'sum',
                '소계' : 'sum'
            })
        )
        end_data.columns = ['항목', '지출소계', '수입소계', '소계']
        end_data.sort_values(by='항목')

        # 데이터 포맷 맞추기
        total_income = mid_data['소계'].sum()
        mid_data['지출소계'] = format.numberFormat(mid_data['지출소계'])
        mid_data['수입소계'] = format.numberFormat(mid_data['수입소계'])
        mid_data['소계'] = format.numberFormat(mid_data['소계'])
        
        # 총 수입 추가
        total_income = format.numberFormat(total_income)
        total_sum = pd.DataFrame(columns=['총 수입'], data=[total_income])
        mid_data = pd.concat([mid_data, total_sum], axis=1)

        end_data['지출소계'] = format.numberFormat(end_data['지출소계'])
        end_data['수입소계'] = format.numberFormat(end_data['수입소계'])
        end_data['소계'] = format.numberFormat(end_data['소계'])

        group['지출금액'] = format.numberFormat(group['지출금액'])
        group['수입금액'] = format.numberFormat(group['수입금액'])
        group['결제일'] = group['결제일'].dt.strftime('%Y-%m-%d')
        group.drop(columns=['월', '소계'], inplace=True)
        
        group.to_excel(writer, sheet_name=month, index=False, startrow=0, startcol=0)
        mid_data.to_excel(writer, sheet_name=month, index=False, startrow=0, startcol=6)
        end_data.to_excel(writer, sheet_name=month, index=False, startrow=0, startcol=12)
        
        resultData = pd.concat([group, mid_data, end_data], axis=1)
        
        # # TODO (UI 개선)디자인 추가
        worksheet = writer.sheets[month]

        # 헤더 색상 적용
        cell_color(worksheet, len(group.columns), 1, 'FFF2CC')
        cell_color(worksheet, len(mid_data.columns), 7, 'CFE2F3')
        cell_color(worksheet, len(end_data.columns), 13, 'F4CCCC')

        # 숫자 오른쪽 정렬
        cell_align(worksheet, [3, 4, 7, 8, 9, 10, 13, 14, 15], 'right')

        # 데이터가 있는 셀은 선 표시
        cell_line(worksheet, dictBorder)

        # 너비 자동 계산
        cell_auto_width(worksheet)
        
